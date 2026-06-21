"""STEP 5+6: K-ESG 자가진단 캐노니컬 출력 + 샘플 회귀.

완료 기준(계획서)
  · K-ESG 28 전 항목이 출력에 존재, 빈 섹션 없음.
  · 각 항목이 verified/self_reported/insufficient/flagged/hitl_required/not_applicable 중 하나.
  · insufficient·hitl 각각 구체적·실행가능한 안내(증빙요청/작성안내)를 단다.
  · 자동응답%·작성필요%·증빙대기% 3분할 정직 보고.
  · 샘플 fixture 회귀 통과(엑셀 출력 누락 없음).
"""
from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from esgenie.knowledge import kesg_items
from esgenie.ssot.audit_trace import DataPoint, EvidenceLink
from esgenie.supplychain import (
    all_framework_keys,
    build_checklist,
    build_response_sheet,
    get_framework,
)
from esgenie.supplychain.exporters import export_response_sheet

_VALID = {
    "verified", "self_reported", "insufficient",
    "flagged", "hitl_required", "not_applicable",
}


# ── STEP 5: 캐노니컬 자가진단 프레임워크 ──────────────────────────────
def test_kesg28_registered_and_complete():
    assert "kesg28" in all_framework_keys()
    fw = get_framework("kesg28")
    assert len(fw.questions) == 28
    # 캐노니컬 코드와 1:1 (빠짐·중복 없음)
    codes = {q.primary_code for q in fw.questions}
    assert codes == set(kesg_items.BASIC_28_CODES)


def test_kesg61_full_profile():
    fw = get_framework("kesg61")
    assert len(fw.questions) == 61


def test_kesg28_no_empty_section():
    sheet = build_response_sheet("kesg28")
    sections = {a.section for a in sheet.answers}
    assert sections == {"정보공시", "환경", "사회", "지배구조"}
    # 각 섹션에 최소 1문항
    for sec in sections:
        assert any(a.section == sec for a in sheet.answers)


def test_every_item_resolves_to_valid_status_empty_input():
    sheet = build_response_sheet("kesg28")
    assert len(sheet.answers) == 28
    for a in sheet.answers:
        assert a.status in _VALID, f"{a.qid}: 잘못된 status {a.status}"


def test_three_way_split_is_reported_and_consistent():
    sheet = build_response_sheet("kesg28")
    # NA 없는 빈입력 → 세 버킷 합 = 100
    assert sheet.denominator == 28
    assert round(sheet.auto_pct + sheet.hitl_pct + sheet.pending_pct, 1) == 100.0
    # 캐노니컬 카탈로그엔 서술필요 항목(S-7-1, G-4-1)이 포함 → 작성필요>0
    assert sheet.hitl_pct > 0.0


def test_insufficient_and_hitl_items_carry_actionable_guidance():
    sheet = build_response_sheet("kesg28")
    for a in sheet.answers:
        if a.status in ("insufficient", "hitl_required"):
            assert a.rationale.strip(), f"{a.qid}: 안내문 비어있음"
            # 정성/정량 미해소엔 올릴문서 또는 작성안내가 있어야 함
            assert a.evidence_needed or a.status == "hitl_required"


# ── STEP 6: 샘플 fixture 회귀 ─────────────────────────────────────────
def _sample_supplier():
    """한울정밀 스타일 협력사 — 일부 정량 증빙 + 일부 정성 공시 보유."""
    mapped_codes = ["E-1-1", "E-1-2", "P-1-1", "S-5-1", "G-1-1", "G-5-1"]
    extraction = SimpleNamespace(
        corp_name="한울정밀",
        mapped={c: {"code": c, "name": c, "evidence_node_ids": []} for c in mapped_codes},
        missing=[],
    )
    data_points = [
        DataPoint(
            kesg_code="E-4-1", kesg_name="에너지 사용량",
            value=128400.0, unit="kWh", period=2025, confidence=0.95,
            verification="verified", d1_risk=0.05,
            evidence_files=[EvidenceLink(
                file_name="한전고지서_2025.pdf", relative_path="evidence_pack/x.pdf",
                origin="ocr_structured", bbox=[0.08, 0.23, 0.3, 0.24], page=0,
                node_id="n1")],
        ),
        DataPoint(
            kesg_code="E-3-1", kesg_name="온실가스(Scope1+2)",
            value=61.39, unit="tCO2eq", period=2025, confidence=0.9,
            verification="estimated", d1_risk=0.12, evidence_files=[],
        ),
    ]
    return extraction, data_points


def test_sample_regression_full_catalog(tmp_path: Path):
    extraction, data_points = _sample_supplier()
    sheet = build_response_sheet(
        "kesg28", corp_name="한울정밀",
        extraction=extraction, data_points=data_points,
    )

    # 1) 빈 섹션 없음 + 전 항목 유효 status
    assert {a.section for a in sheet.answers} == {"정보공시", "환경", "사회", "지배구조"}
    assert all(a.status in _VALID for a in sheet.answers)

    # 2) 실제로 verified/self_reported 가 생겨 자동응답% > 0
    statuses = {a.status for a in sheet.answers}
    assert "verified" in statuses or "self_reported" in statuses
    assert sheet.auto_pct > 0.0

    # 3) 에너지(E-4-1) 정량 증빙 → verified
    energy = next(a for a in sheet.answers if a.qid == "KESG-E-4-1")
    assert energy.value == 128400.0
    assert energy.status == "verified"

    # 4) 체크리스트엔 해소된 항목이 빠진다
    checklist_qids = {it.qid for it in build_checklist(sheet)}
    assert "KESG-E-4-1" not in checklist_qids
    assert "KESG-E-2-1" in checklist_qids  # 원부자재 미제공 → 남음

    # 5) 엑셀 출력 누락 없음
    path = export_response_sheet(sheet, tmp_path)
    wb = load_workbook(path)
    assert "응답서" in wb.sheetnames
    assert "증빙 체크리스트" in wb.sheetnames
    ws = wb["응답서"]
    # 헤더(4행) 다음부터 28개 데이터 행
    data_rows = [r for r in ws.iter_rows(min_row=5, max_col=1, values_only=True) if r[0]]
    assert len(data_rows) == 28
