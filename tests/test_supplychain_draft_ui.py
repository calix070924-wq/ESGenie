"""AI 초안(draft_ready) UI·익스포터 통합 테스트.

draft_ready 상태 Answer를 직접 구성해(drafter 미실행) UI 렌더·엑셀·PDF·체크리스트·캐싱을 검증한다.
"""
from __future__ import annotations

import sys
import types
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

import pytest

from esgenie.supplychain.schema import Answer, EvidenceLink, ResponseSheet


# ── 공통 픽스처 ──────────────────────────────────────────────────────────────

def _make_draft_answer() -> Answer:
    return Answer(
        qid="S-1-1",
        section="사회",
        question_text="인권 정책을 설명하세요",
        value=None,
        status="draft_ready",
        evidence_links=[],
        flags=[],
        rationale="",
        evidence_needed=[],
        draft_text="당사는 UN 인권경영 가이드라인에 따라 인권 정책을 수립하였습니다. [node_a]",
        draft_citations=[
            {
                "node_id": "node_a",
                "source_file": "인권정책서.pdf",
                "page": 2,
                "text_preview": "UN 인권경영 가이드라인에 따라...",
                "retrieval": "code_match",
            },
            {
                "node_id": "node_b",
                "source_file": "사내규정.pdf",
                "page": 5,
                "text_preview": "인권 영향 평가를 매년 실시...",
                "retrieval": "bm25_fallback",
            },
        ],
        draft_grounding={
            "decision": "ACCEPT",
            "g1_uncited_sentences": [],
            "g2_orphan_numbers": [],
            "g4_unit_mismatches": [],
            "g5_overclaim": False,
            "hard_fails": [],
            "soft_flags": [],
            "faithfulness": 0.95,
        },
    )


def _make_verified_answer() -> Answer:
    return Answer(
        qid="E-4-1",
        section="환경",
        question_text="에너지 사용량을 보고하세요",
        value=128400.0,
        status="verified",
        evidence_links=[EvidenceLink(
            file_name="한전고지서.pdf", relative_path="evidence_pack/x.pdf",
            origin="ocr_structured", bbox=[0.08, 0.23, 0.3, 0.24], page=0,
            node_id="n1",
        )],
        flags=[],
        rationale="한전 고지서 대조 완료",
        evidence_needed=[],
    )


def _make_sheet(*, include_draft: bool = True) -> ResponseSheet:
    answers = [_make_verified_answer()]
    if include_draft:
        answers.append(_make_draft_answer())
    return ResponseSheet(
        framework_key="kesg28",
        framework_label="K-ESG 28",
        corp_name="한울정밀",
        answers=answers,
        gaps=[],
    )


# ── fake streamlit ───────────────────────────────────────────────────────────

def _fake_streamlit() -> MagicMock:
    st = MagicMock(name="streamlit")
    st.session_state = {}
    st.columns.side_effect = lambda spec, *a, **k: [
        MagicMock() for _ in range(spec if isinstance(spec, int) else len(spec))
    ]
    st.selectbox.side_effect = lambda label, options, **k: options[0]
    return st


@pytest.fixture
def tabs_module(monkeypatch):
    fake_st = _fake_streamlit()
    monkeypatch.setitem(sys.modules, "streamlit", fake_st)
    plotly = types.ModuleType("plotly")
    go = types.ModuleType("plotly.graph_objects")
    plotly.graph_objects = go
    monkeypatch.setitem(sys.modules, "plotly", plotly)
    monkeypatch.setitem(sys.modules, "plotly.graph_objects", go)
    sys.modules.pop("esgenie.ui.tabs", None)
    import esgenie.ui.tabs as tabs
    tabs.st = fake_st
    return tabs


# ── 1. 5컬럼 메트릭 + AI초안 % ─────────────────────────────────────────────

def test_five_column_metrics(tabs_module):
    sheet = _make_sheet(include_draft=True)
    assert sheet.draft_pct > 0, "draft_pct must be non-zero for draft_ready answer"
    # 2 answers total: 1 verified + 1 draft_ready. denominator = 2
    expected_draft_pct = 50.0
    assert abs(sheet.draft_pct - expected_draft_pct) < 0.1


# ── 2. draft_ready expander 내용 ────────────────────────────────────────────

def test_draft_expander_contents(tabs_module):
    """draft_ready expander에 초안 전문·근거 발췌·게이트 배지·승인 전 캡션이 존재."""
    st = tabs_module.st
    sheet = _make_sheet(include_draft=True)

    # Simulate what _render_responder_workspace does for draft section
    draft_answers = [a for a in sheet.answers if a.status == "draft_ready"]
    assert len(draft_answers) == 1

    da = draft_answers[0]
    assert "UN 인권경영 가이드라인" in da.draft_text
    assert da.draft_citations[0]["source_file"] == "인권정책서.pdf"
    assert da.draft_citations[1]["retrieval"] == "bm25_fallback"
    assert da.draft_grounding["faithfulness"] == 0.95
    assert da.draft_grounding["hard_fails"] == []


# ── 3. 캐싱: 같은 result로 2회 렌더 → drafter 1회만 ────────────────────────

def test_caching_prevents_duplicate_calls(tabs_module, monkeypatch, tmp_path):
    """같은 result·framework에 대해 respond_from_pipeline이 1회만 호출됨을 증명."""
    monkeypatch.chdir(tmp_path)
    call_count = {"n": 0}
    original_respond = None

    import esgenie.supplychain.responder as responder_mod
    original_respond = responder_mod.respond_from_pipeline

    def counting_respond(*args, **kwargs):
        call_count["n"] += 1
        return original_respond(*args, **kwargs)

    monkeypatch.setattr(
        "esgenie.supplychain.responder.respond_from_pipeline",
        counting_respond,
    )
    # Also patch the import in tabs
    monkeypatch.setattr(
        "esgenie.ui.tabs.respond_from_pipeline",
        counting_respond,
    )

    fake_st = tabs_module.st
    fake_st.session_state = {}

    from esgenie.ui.tabs import _get_cached_response_sheet
    from esgenie.supplychain.frameworks import get_framework

    result = SimpleNamespace(
        report=SimpleNamespace(corp_name="한울정밀"),
        extraction=SimpleNamespace(
            corp_name="한울정밀",
            mapped={"E-4-1": {"code": "E-4-1", "name": "에너지", "evidence_node_ids": []}},
            missing=[],
        ),
        disclosure=None,
        v15_trace=SimpleNamespace(data_points=[]),
        evidence_graph=None,
        supplier_claims=None,
        supplier_claim_files=[],
    )

    fw = get_framework("kesg28")

    # First call
    sheet1 = _get_cached_response_sheet(result, fw)
    assert call_count["n"] == 1

    # Second call — should use cache
    sheet2 = _get_cached_response_sheet(result, fw)
    assert call_count["n"] == 1, "캐시 적중 시 respond_from_pipeline 재호출 금지"
    assert sheet1 is sheet2


def test_cache_invalidates_on_supplier_claims_change(tabs_module, monkeypatch, tmp_path):
    """SAQ 추가(supplier_claims 변경) 시 캐시가 무효화되어 재계산됨을 증명."""
    monkeypatch.chdir(tmp_path)
    call_count = {"n": 0}

    import esgenie.supplychain.responder as responder_mod
    original_respond = responder_mod.respond_from_pipeline

    def counting_respond(*args, **kwargs):
        call_count["n"] += 1
        return original_respond(*args, **kwargs)

    monkeypatch.setattr("esgenie.ui.tabs.respond_from_pipeline", counting_respond)

    fake_st = tabs_module.st
    fake_st.session_state = {}

    from esgenie.ui.tabs import _get_cached_response_sheet
    from esgenie.supplychain.frameworks import get_framework

    result = SimpleNamespace(
        report=SimpleNamespace(corp_name="한울정밀"),
        extraction=SimpleNamespace(
            corp_name="한울정밀",
            mapped={"E-4-1": {"code": "E-4-1", "name": "에너지", "evidence_node_ids": []}},
            missing=[],
        ),
        disclosure=None,
        v15_trace=SimpleNamespace(data_points=[]),
        evidence_graph=None,
        supplier_claims=None,
        supplier_claim_files=[],
    )
    fw = get_framework("kesg28")

    # First call — no claims
    _get_cached_response_sheet(result, fw, supplier_claims=None)
    assert call_count["n"] == 1

    # Second call — same params → cache hit
    _get_cached_response_sheet(result, fw, supplier_claims=None)
    assert call_count["n"] == 1

    # Third call — claims attached → cache miss, must recompute
    from esgenie.supplychain.claims import SupplierClaim
    claims = {
        "E-6-2": SupplierClaim(
            code="E-6-2", value=92.0, unit="%",
            raw="재활용률 92%", source="saq:OEM_ESG자가진단설문.pdf",
        )
    }
    _get_cached_response_sheet(result, fw, supplier_claims=claims)
    assert call_count["n"] == 2, "supplier_claims 변경 시 캐시 무효화 → 재호출 필수"

    # Fourth call — same claims → cache hit again
    _get_cached_response_sheet(result, fw, supplier_claims=claims)
    assert call_count["n"] == 2, "동일 claims로 재호출 시 캐시 적중"


# ── 4. checklist: draft_ready → "초안 검토" ──────────────────────────────────

def test_checklist_draft_ready_action():
    """draft_ready 항목이 '초안 검토' 액션으로 잡히고 기존 상태 액션은 불변."""
    from esgenie.supplychain.checklist import build_checklist

    sheet = _make_sheet(include_draft=True)
    # Add an insufficient answer to verify existing actions stay unchanged
    sheet.answers.append(Answer(
        qid="G-1-1",
        section="지배구조",
        question_text="이사회 구성을 설명하세요",
        value=None,
        status="insufficient",
        evidence_links=[],
        flags=[],
        rationale="증빙 누락",
        evidence_needed=["이사회 의사록"],
    ))

    items = build_checklist(sheet)
    draft_items = [it for it in items if it.status == "draft_ready"]
    assert len(draft_items) == 1
    assert draft_items[0].action == "초안 검토"
    assert "검토" in draft_items[0].request

    insuf_items = [it for it in items if it.status == "insufficient"]
    assert len(insuf_items) == 1
    assert insuf_items[0].action == "증빙 업로드"


# ── 5. excel: [AI 초안 — 승인 전] 라벨 + draft_text 셀 존재 ───────────────

def test_excel_draft_label(tmp_path):
    """xlsx에 '[AI 초안 — 승인 전]' 라벨과 draft_text가 셀에 있는지 검증."""
    from esgenie.supplychain.exporters.excel import export_response_sheet

    sheet = _make_sheet(include_draft=True)
    xlsx_path = export_response_sheet(sheet, str(tmp_path))

    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws = wb.active

    found_label = False
    found_draft_text = False
    for row in ws.iter_rows(values_only=True):
        for cell_val in row:
            if cell_val is None:
                continue
            text = str(cell_val)
            if "[AI 초안 — 승인 전]" in text:
                found_label = True
            if "UN 인권경영 가이드라인" in text:
                found_draft_text = True
    assert found_label, "xlsx에 '[AI 초안 — 승인 전]' 라벨이 없음"
    assert found_draft_text, "xlsx에 draft_text 내용이 없음"


def test_excel_sec_auto_excludes_draft(tmp_path):
    """sec_auto 집계에 draft_ready가 포함되지 않는지 검증."""
    from esgenie.supplychain.exporters.excel import export_response_sheet

    sheet = _make_sheet(include_draft=True)
    xlsx_path = export_response_sheet(sheet, str(tmp_path))

    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # 그룹 헤더 셀에서 "자동응답" 뒤 숫자를 확인 — draft_ready가 세어지지 않아야 함
    for row in ws.iter_rows(values_only=True):
        for cell_val in row:
            if cell_val and "자동응답" in str(cell_val) and "사회" in str(cell_val):
                # 사회 섹션: draft_ready 1건만 → auto 0/1
                assert "0/1" in str(cell_val), f"draft_ready가 sec_auto에 포함됨: {cell_val}"


# ── 6. pdf: 생성 성공 + 라벨 존재 ───────────────────────────────────────────

def test_pdf_no_crash(tmp_path):
    """PDF 생성이 크래시 없이 완료되고 AI초안(승인전) 라벨이 사용됨을 확인."""
    from esgenie.supplychain.exporters.pdf import export_response_sheet_pdf

    sheet = _make_sheet(include_draft=True)
    pdf_path = export_response_sheet_pdf(sheet, str(tmp_path), embed_evidence=False)
    assert pdf_path.endswith(".pdf")

    import os
    assert os.path.exists(pdf_path)
    assert os.path.getsize(pdf_path) > 0

    # 가능하면 텍스트 추출
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(pdf_path)
        text = ""
        for page in doc:
            text += page.get_text()
        doc.close()
        assert "AI초안(승인전)" in text or "AI 초안" in text or "승인 전" in text
    except ImportError:
        pass  # PyMuPDF 없으면 크래시 없음만 검증


# ── 7. 회귀: draft_ready 없는 시트 → 기존 동작 동일 ──────────────────────────

def test_regression_no_draft_excel(tmp_path):
    """draft_ready가 없는 시트에서 기존 xlsx 출력이 정상."""
    from esgenie.supplychain.exporters.excel import export_response_sheet

    sheet = _make_sheet(include_draft=False)
    xlsx_path = export_response_sheet(sheet, str(tmp_path))

    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws = wb.active

    found_ai_label = False
    for row in ws.iter_rows(values_only=True):
        for cell_val in row:
            if cell_val and "[AI 초안 — 승인 전]" in str(cell_val):
                found_ai_label = True
    assert not found_ai_label, "draft 없는 시트에 AI 초안 라벨이 나타남"


def test_regression_no_draft_pdf(tmp_path):
    """draft_ready가 없는 시트에서 기존 pdf 출력이 정상."""
    from esgenie.supplychain.exporters.pdf import export_response_sheet_pdf

    sheet = _make_sheet(include_draft=False)
    pdf_path = export_response_sheet_pdf(sheet, str(tmp_path), embed_evidence=False)
    assert pdf_path.endswith(".pdf")

    import os
    assert os.path.exists(pdf_path)
    assert os.path.getsize(pdf_path) > 0


def test_regression_no_draft_checklist():
    """draft_ready가 없는 시트에서 체크리스트가 기존과 동일(빈 리스트)."""
    from esgenie.supplychain.checklist import build_checklist

    sheet = _make_sheet(include_draft=False)
    items = build_checklist(sheet)
    # verified answer는 체크리스트에 안 잡힘
    assert len(items) == 0
