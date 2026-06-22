from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from esgenie.ssot.ocr_router import DocChannel, ExtractedTable, OcrExtraction, TableCell
from esgenie.ssot.table_benchmark import (
    build_review_workbook,
    extract_table_cases,
    load_cases_jsonl,
    make_case_id,
    write_cases_jsonl,
)


def _table() -> ExtractedTable:
    return ExtractedTable(
        table_id="azure_table_0",
        row_count=4,
        column_count=2,
        page=0,
        meta={"resolved_tier": 1, "gate_flags": ["A2_blank_ratio"], "hard_fail_codes": []},
        cells=[
            TableCell(0, 0, "항목", confidence=0.99),
            TableCell(0, 1, "2025 (kWh)", confidence=0.99),
            TableCell(1, 0, "전력", confidence=0.95),
            TableCell(1, 1, "100", confidence=0.95),
            TableCell(2, 0, "가스", confidence=0.95),
            TableCell(2, 1, "50", confidence=0.95),
            TableCell(3, 0, "합계", confidence=0.95),
            TableCell(3, 1, "150", confidence=0.95),
        ],
    )


def test_make_case_id():
    assert make_case_id("01_전기요금청구서_2026-05.pdf", 0, "azure_table_0").startswith("01_전기요금청구서_2026_05__p1")


def test_extract_table_cases_writes_snapshot(tmp_path: Path):
    ext = OcrExtraction(
        source_file="01_전기요금청구서_2026-05.pdf",
        channel=DocChannel.STRUCTURED,
        doc_type="kepco_bill",
        tables=[_table()],
        router_meta={
            "table_gate": {
                "status": "ACCEPT",
                "resolved_tier": 1,
                "needs_semantic_check": False,
                "tables": [{
                    "table_id": "azure_table_0",
                    "decision": "ACCEPT",
                    "hard_fail_count": 0,
                    "soft_flag_count": 1,
                    "flags": [],
                }],
            }
        },
    )
    cases, summary = extract_table_cases(
        [ext],
        split="dev",
        raw_dir=tmp_path / "raw",
        source_paths={ext.source_file: "/abs/01_전기요금청구서_2026-05.pdf"},
    )
    assert summary["n_cases"] == 1
    assert cases[0]["gate_status_pred"] == "ACCEPT"
    assert cases[0]["resolved_tier_pred"] == 1
    assert cases[0]["snapshot_path"].startswith("raw/")
    snap = tmp_path / "raw" / f"{cases[0]['id']}.json"
    assert snap.exists()
    payload = json.loads(snap.read_text(encoding="utf-8"))
    assert payload["table"]["table_id"] == "azure_table_0"


def test_write_cases_jsonl_and_load_roundtrip(tmp_path: Path):
    cases = [{
        "id": "c1",
        "split": "dev",
        "source_file": "a.pdf",
        "source_path": "/abs/a.pdf",
        "doc_type": "kepco_bill",
        "page": 1,
        "table_id": "t0",
        "gate_status_pred": "ACCEPT",
        "resolved_tier_pred": 1,
        "needs_semantic_check_pred": False,
        "hard_fail_codes": "",
        "soft_flag_codes": "A2_blank_ratio",
        "error_tags": "A2_blank_ratio",
        "n_rows": 4,
        "n_cols": 2,
        "snapshot_path": "raw/c1.json",
        "gold_table_path": "data/benchmark_tables/gold/c1.csv",
        "gate_gold": "",
        "review_status": "pending",
        "notes": "",
    }]
    path = write_cases_jsonl(cases, tmp_path / "cases.jsonl")
    loaded = load_cases_jsonl(path)
    assert loaded == cases


def test_build_review_workbook(tmp_path: Path):
    cases = [{
        "id": "c1",
        "split": "dev",
        "source_file": "a.pdf",
        "source_path": "/abs/a.pdf",
        "doc_type": "kepco_bill",
        "page": 1,
        "table_id": "t0",
        "gate_status_pred": "ACCEPT",
        "resolved_tier_pred": 1,
        "needs_semantic_check_pred": False,
        "hard_fail_codes": "",
        "soft_flag_codes": "A2_blank_ratio",
        "error_tags": "A2_blank_ratio",
        "n_rows": 4,
        "n_cols": 2,
        "snapshot_path": "raw/c1.json",
        "gold_table_path": "data/benchmark_tables/gold/c1.csv",
        "gate_gold": "",
        "review_status": "pending",
        "notes": "",
    }]
    out = build_review_workbook(cases, tmp_path / "review.xlsx")
    wb = load_workbook(out)
    assert {"Cases", "Summary", "Guide"} <= set(wb.sheetnames)
    ws = wb["Cases"]
    assert ws["A2"].value == "c1"
    assert ws["H2"].value == "ACCEPT"
    assert ws["R2"].value is None
