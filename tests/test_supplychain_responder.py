"""공급망 실사 응답 모듈 PR1 — 골격 동작 테스트.

불변식:
  1) 입력 없음 → 모든 문항 '데이터부족'(insufficient), 커버리지 0 (회귀/빈입력 안전)
  2) 추출 + 증빙 data_point 주어지면 → 해당 칸이 채워지고 verified/self_reported
  3) D6 고아 비율(폐기물 재활용률 분모 누락) → 해당 답변 flagged + D6 플래그
"""
from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from esgenie.issb_gap import build_issb_gap_report
from esgenie.layer3_disclosure import DisclosureReport, OrphanRatio
from esgenie.ssot.audit_trace import DataPoint, EvidenceLink
from esgenie.ssot.evidence_graph import EvidenceGraph, TextNode
from esgenie.supplychain import (
    all_framework_keys,
    build_response_sheet,
    get_framework,
    manual_claims,
)
from esgenie.supplychain.exporters import export_response_sheet

FW = "saq5_env"


def _extraction(mapped_codes, missing=()):
    return SimpleNamespace(
        corp_name="한국정밀",
        profile="full",
        profile_label="전체 (61항목)",
        mapped={c: {"code": c, "name": c, "evidence_node_ids": []} for c in mapped_codes},
        missing=list(missing),
    )


def _energy_datapoint():
    return DataPoint(
        kesg_code="E-4-1", kesg_name="에너지 사용량",
        value=128400.0, unit="kWh", period=2025, confidence=0.95,
        verification="verified", d1_risk=0.05,
        evidence_files=[EvidenceLink(
            file_name="한전고지서_2025_03.pdf",
            relative_path="evidence_pack/한전고지서_2025_03.pdf",
            origin="ocr_structured", bbox=[0.084, 0.234, 0.306, 0.247],
            page=0, node_id="ocr_E-4-1_2025",
        )],
    )


def _ghg_datapoint():
    return DataPoint(
        kesg_code="E-3-1", kesg_name="온실가스 배출량(Scope1+2)",
        value=61.39, unit="tCO2eq", period=2025, confidence=0.91,
        verification="estimated", d1_risk=0.12,
        evidence_files=[EvidenceLink(
            file_name="한전고지서_2025_03.pdf",
            relative_path="evidence_pack/한전고지서_2025_03.pdf",
            origin="ocr_structured", bbox=[0.084, 0.234, 0.306, 0.247],
            page=0, node_id="LOCAL_E-3-1_2025__derived_ocr_structured",
        )],
    )


# ── 1. 회귀/빈입력 가드 ────────────────────────────────────────────
def test_empty_input_all_insufficient():
    sheet = build_response_sheet(FW)
    assert sheet.coverage_pct == 0.0
    assert sheet.answers, "문항이 비어있으면 안 됨"
    assert all(a.status == "insufficient" for a in sheet.answers)
    # 보완 안내가 문항 수만큼 생성됨
    assert len(sheet.gaps) >= len(sheet.answers)


def test_framework_registry():
    assert FW in all_framework_keys()
    assert get_framework(FW).questions
    assert "saq5" in all_framework_keys()
    full = get_framework("saq5")
    assert len({q.section for q in full.questions}) >= 5
    assert any(q.qid == "SAQ-E-NUM-GHG" for q in full.questions)
    with pytest.raises(KeyError):
        get_framework("does_not_exist")


# ── 2. 추출 + 증빙 → 응답 채움 ─────────────────────────────────────
def test_populates_from_extraction_and_datapoint():
    extraction = _extraction(
        mapped_codes=["E-1-1", "E-1-2", "E-3-1", "E-4-1", "E-6-2"],
    )
    sheet = build_response_sheet(
        FW, extraction=extraction, data_points=[_ghg_datapoint(), _energy_datapoint()],
    )
    by_id = {a.qid: a for a in sheet.answers}

    # 환경방침 보유(E-1-1/E-1-2 공시) → 예
    assert by_id["SAQ-E-10"].value is True

    # Scope 1+2 배출량 수치 문항 지원
    ghg = by_id["SAQ-E-NUM-GHG"]
    assert ghg.value == 61.39
    assert ghg.status == "self_reported"
    assert ghg.evidence_links and ghg.evidence_links[0].bbox is not None

    # 에너지 사용량 → 증빙 data_point로 검증됨
    energy = by_id["SAQ-E-NUM-ENERGY"]
    assert energy.value == 128400.0
    assert energy.status == "verified"
    assert energy.evidence_links and energy.evidence_links[0].bbox is not None

    # 커버리지가 0보다 큼
    assert sheet.coverage_pct > 0.0


def test_multi_select_partial_coverage():
    extraction = _extraction(mapped_codes=["E-4-1", "E-6-1"])  # 2개 영역만
    sheet = build_response_sheet(FW, extraction=extraction)
    q10a = next(a for a in sheet.answers if a.qid == "SAQ-E-10a")
    assert "에너지 효율" in q10a.value
    assert "폐기물 감축" in q10a.value
    assert "재생에너지" not in q10a.value          # 미공시
    assert any(f.startswith("미충족") for f in q10a.flags)


def test_presence_answer_resolves_file_link_from_text_node():
    graph = EvidenceGraph("SME001", "한국정밀")
    tnode = TextNode(
        id="SME001_TXT_0001",
        section="환경방침",
        text="당사는 환경방침을 수립하고 전사에 배포한다.",
        kesg_code="E-1-1",
        source_file="환경방침서.pdf",
        page=2,
    )
    graph.add_text_node(tnode)
    extraction = _extraction(mapped_codes=[])
    extraction.mapped["E-1-1"] = {
        "code": "E-1-1",
        "name": "환경경영 목표 수립",
        "area": "E",
        "category": "환경경영 목표",
        "data_type": "정성",
        "value": "문서 조항 확인",
        "unit": "",
        "note": "OCR 정성 증빙",
        "evidence_node_ids": [tnode.id],
        "beyond_profile": False,
    }

    sheet = build_response_sheet(FW, extraction=extraction, evidence_graph=graph)
    answer = next(a for a in sheet.answers if a.qid == "SAQ-E-10")
    assert answer.evidence_links
    assert answer.evidence_links[0].file_name == "환경방침서.pdf"
    assert answer.evidence_links[0].page == 2
    assert answer.status == "verified"


# ── 3. D6 게이팅: 고아 비율 → flagged ──────────────────────────────
def test_d6_orphan_ratio_flags_waste_recycling():
    extraction = _extraction(mapped_codes=["E-6-2"], missing=["E-6-1"])
    disclosure = DisclosureReport(
        score=0.6, level="high",
        orphan_ratios=[OrphanRatio(
            ratio_code="E-6-2", ratio_name="폐기물 재활용 비율",
            missing_context=["E-6-1"],
            detail="'폐기물 재활용 비율'은 공시하면서 맥락 항목(폐기물 배출량) 누락",
        )],
    )
    sheet = build_response_sheet(FW, extraction=extraction, disclosure=disclosure)
    waste = next(a for a in sheet.answers if a.qid == "SAQ-E-NUM-WASTE")
    assert waste.status == "flagged"
    assert any("D6" in f for f in waste.flags)
    assert sheet.flagged_count >= 1
    assert any("검토" in g for g in sheet.gaps)


def test_no_disclosure_no_crash():
    extraction = _extraction(mapped_codes=["E-6-2"])
    sheet = build_response_sheet(FW, extraction=extraction, disclosure=None)
    assert sheet.answers  # disclosure None이어도 동작


def test_supplier_claim_treats_e62_as_ratio_even_if_unit_string_is_wrong():
    extraction = _extraction(mapped_codes=[])
    extraction.mapped["E-6-2"] = {
        "code": "E-6-2",
        "name": "폐기물 재활용 비율",
        "value": 29.3,
        "unit": "ton",
        "evidence_node_ids": [],
    }

    sheet = build_response_sheet(
        FW,
        extraction=extraction,
        supplier_claims=manual_claims({"E-6-2": 29.3}),
    )
    waste = next(a for a in sheet.answers if a.qid == "SAQ-E-NUM-WASTE")
    assert waste.status == "self_reported"
    assert any("자가신고 일치" in f for f in waste.flags)
    assert all("비율(%) 미확보" not in f for f in waste.flags)


def test_supplier_claim_flags_out_of_range_value_for_ratio_code():
    extraction = _extraction(mapped_codes=[])
    extraction.mapped["E-6-2"] = {
        "code": "E-6-2",
        "name": "폐기물 재활용 비율",
        "value": 5400.0,
        "unit": "ton",
        "evidence_node_ids": [],
    }

    sheet = build_response_sheet(
        FW,
        extraction=extraction,
        supplier_claims=manual_claims({"E-6-2": 29.3}),
    )
    waste = next(a for a in sheet.answers if a.qid == "SAQ-E-NUM-WASTE")
    assert waste.status == "flagged"
    assert any("비율 범위(0~100%)" in f for f in waste.flags)


def test_issb_missing_climate_item_flags_supplychain_answer():
    extraction = _extraction(mapped_codes=["E-1-1"], missing=["E-4-1"])
    issb_gap = build_issb_gap_report(extraction)
    sheet = build_response_sheet(FW, extraction=extraction, issb_gap=issb_gap)
    energy = next(a for a in sheet.answers if a.qid == "SAQ-E-NUM-ENERGY")
    assert energy.status == "flagged"
    assert any("ISSB S2" in f for f in energy.flags)
    assert any("보완 증빙:" in f for f in energy.flags)
    assert "전기·가스 사용량 집계표" in energy.rationale
    assert any("검토" in g and "ISSB" in g for g in sheet.gaps)


def test_issb_gap_none_no_crash():
    extraction = _extraction(mapped_codes=["E-4-1"])
    sheet = build_response_sheet(FW, extraction=extraction, disclosure=None, issb_gap=None)
    assert sheet.answers


# ── 4. Excel 출력 ──────────────────────────────────────────────────
def test_excel_export(tmp_path: Path):
    extraction = _extraction(mapped_codes=["E-1-1", "E-4-1", "E-6-2"], missing=["E-6-1"])
    disclosure = DisclosureReport(
        score=0.6, level="high",
        orphan_ratios=[OrphanRatio(
            ratio_code="E-6-2", ratio_name="폐기물 재활용 비율",
            missing_context=["E-6-1"], detail="분모 누락",
        )],
    )
    sheet = build_response_sheet(
        FW, corp_name="한국정밀", extraction=extraction,
        disclosure=disclosure, data_points=[_energy_datapoint()],
    )
    path = export_response_sheet(sheet, tmp_path)
    assert Path(path).exists()
    assert Path(path).suffix == ".xlsx"


def test_excel_export_adds_issb_followup_sheet(tmp_path: Path):
    extraction = _extraction(mapped_codes=["E-1-1"], missing=["E-4-1"])
    issb_gap = build_issb_gap_report(extraction)
    sheet = build_response_sheet(FW, corp_name="한국정밀", extraction=extraction, issb_gap=issb_gap)

    path = export_response_sheet(sheet, tmp_path)
    wb = load_workbook(path)

    assert "ISSB 보완" in wb.sheetnames
    ws = wb["ISSB 보완"]
    assert ws["A1"].value == "ISSB/KSSB 보완 항목"
    rows = list(ws.iter_rows(min_row=5, values_only=True))
    assert any(row[0] == "(수치) 연간 에너지 사용량" for row in rows)
    energy_row = next(row for row in rows if row[0] == "(수치) 연간 에너지 사용량")
    assert "ISSB S2" in str(energy_row[1])
    assert "전기·가스 사용량 집계표" in str(energy_row[2])


def test_excel_export_skips_issb_followup_sheet_when_no_issb_flags(tmp_path: Path):
    extraction = _extraction(mapped_codes=["E-1-1", "E-3-1", "E-4-1"], missing=[])
    sheet = build_response_sheet(FW, corp_name="한국정밀", extraction=extraction)

    path = export_response_sheet(sheet, tmp_path)
    wb = load_workbook(path)

    assert "ISSB 보완" not in wb.sheetnames


# ── 5. STEP 1: status 모델 확장 + 3분할 집계 ───────────────────────────
from esgenie.supplychain.schema import Answer, ResponseSheet  # noqa: E402


def _answer(qid: str, status: str) -> Answer:
    return Answer(qid=qid, section="X", question_text=qid, value=None, status=status)


def test_new_statuses_have_badges():
    # hitl_required / not_applicable 도 배지 매핑이 있어야 함(렌더 시 KeyError 금지)
    assert _answer("a", "hitl_required").badge == "✍️ 작성필요"
    assert _answer("b", "not_applicable").badge == "➖ 해당없음"


def test_answered_excludes_hitl_and_na():
    # 자동응답으로 카운트되는 건 verified/self_reported/flagged 뿐
    assert _answer("a", "verified").answered
    assert _answer("a", "self_reported").answered
    assert _answer("a", "flagged").answered
    assert not _answer("a", "hitl_required").answered
    assert not _answer("a", "insufficient").answered
    assert not _answer("a", "not_applicable").answered


def test_coverage_three_way_split_sums_to_100():
    sheet = ResponseSheet(
        framework_key="t", framework_label="t", corp_name="c",
        answers=[
            _answer("1", "verified"),       # 자동
            _answer("2", "self_reported"),  # 자동
            _answer("3", "flagged"),        # 자동
            _answer("4", "hitl_required"),  # 사람필요
            _answer("5", "insufficient"),   # 증빙대기
        ],
    )
    assert sheet.denominator == 5
    assert sheet.auto_pct == 60.0
    assert sheet.hitl_pct == 20.0
    assert sheet.pending_pct == 20.0
    assert sheet.auto_pct + sheet.hitl_pct + sheet.pending_pct == 100.0
    # 하위호환: coverage_pct == auto_pct
    assert sheet.coverage_pct == sheet.auto_pct


def test_not_applicable_excluded_from_denominator():
    sheet = ResponseSheet(
        framework_key="t", framework_label="t", corp_name="c",
        answers=[
            _answer("1", "verified"),
            _answer("2", "not_applicable"),  # 분모에서 빠짐
            _answer("3", "not_applicable"),
        ],
    )
    assert sheet.denominator == 1
    assert sheet.auto_pct == 100.0  # 1/1, NA 제외


def test_all_not_applicable_no_zero_division():
    sheet = ResponseSheet(
        framework_key="t", framework_label="t", corp_name="c",
        answers=[_answer("1", "not_applicable")],
    )
    assert sheet.denominator == 0
    assert sheet.auto_pct == 0.0
    assert sheet.hitl_pct == 0.0
    assert sheet.pending_pct == 0.0


def test_to_dict_includes_three_way_split():
    sheet = ResponseSheet(
        framework_key="t", framework_label="t", corp_name="c",
        answers=[_answer("1", "verified"), _answer("2", "hitl_required")],
    )
    d = sheet.to_dict()
    assert d["auto_pct"] == 50.0
    assert d["hitl_pct"] == 50.0
    assert d["pending_pct"] == 0.0


def test_excel_export_handles_new_statuses(tmp_path: Path):
    # hitl_required/not_applicable 행이 섞여도 색칠·저장에서 KeyError 없이 동작
    sheet = ResponseSheet(
        framework_key="t", framework_label="테스트", corp_name="한국정밀",
        answers=[
            _answer("1", "verified"),
            _answer("2", "hitl_required"),
            _answer("3", "not_applicable"),
        ],
    )
    path = export_response_sheet(sheet, tmp_path)
    assert Path(path).exists()


# ── 6. STEP 3: derive를 데이터타입 기준으로 라우팅 ──────────────────────
def test_empty_full_saq5_routes_narrative_item_to_hitl():
    # 전체본(saq5)은 G-4-1(윤리위반 공시, 서술필요) 문항을 포함 → 빈입력 시 hitl_required
    sheet = build_response_sheet("saq5")
    ethics = next(a for a in sheet.answers if a.qid == "SAQ-B-1")
    assert ethics.status == "hitl_required"
    assert ethics.badge == "✍️ 작성필요"
    # 작성필요 버킷이 0이 아니어야 함(3분할이 실제로 갈림)
    assert sheet.hitl_pct > 0.0


def test_empty_full_saq5_evidence_policy_item_stays_insufficient():
    # 정보보호(S-8-1)는 정성이지만 증빙 올리면 풀리는 항목 → hitl 아닌 insufficient
    sheet = build_response_sheet("saq5")
    infosec = next(a for a in sheet.answers if a.qid == "SAQ-B-2")
    assert infosec.status == "insufficient"
    assert "정보보호" in infosec.rationale  # 룩업 테이블의 구체 안내문이 실림


def test_unresolved_quantitative_uses_specific_request():
    # 미해소 정량 문항은 insufficient + 룩업의 구체 증빙요청
    sheet = build_response_sheet(FW)  # saq5_env, 빈입력
    energy = next(a for a in sheet.answers if a.qid == "SAQ-E-NUM-ENERGY")
    assert energy.status == "insufficient"
    assert "고지서" in energy.rationale  # E-4-1 request에 고지서 언급


def test_env_only_framework_has_no_hitl_when_empty():
    # saq5_env는 E 코드만 → 빈입력 시 서술필요 항목이 없어 전부 insufficient(STEP1 회귀 유지)
    sheet = build_response_sheet(FW)
    assert all(a.status == "insufficient" for a in sheet.answers)
    assert sheet.hitl_pct == 0.0


def test_gaps_surface_hitl_as_write_task():
    sheet = build_response_sheet("saq5")
    assert any(g.startswith("[작성]") for g in sheet.gaps)
