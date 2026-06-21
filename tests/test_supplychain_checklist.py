"""STEP 4: 수직 슬라이스 관통 — 노동·인권 섹션 증빙 체크리스트.

증명 목표(계획서 STEP 4): 안 풀리는 섹션 하나를 골라 전 경로 끝까지 도는가.
  derive → (status·evidence_needed 적재) → checklist → exporter(excel) → UI rows
"""
from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from esgenie.supplychain import (
    build_checklist,
    build_response_sheet,
    checklist_rows,
)
from esgenie.supplychain.exporters import export_response_sheet

SECTION = "Labor & Human Rights"  # saq5: SAQ-L-1(S-1-1)/L-2(S-2-6)/L-3(S-5-1)


def _extraction(mapped_codes, missing=()):
    return SimpleNamespace(
        corp_name="한울정밀",
        mapped={c: {"code": c, "name": c, "evidence_node_ids": []} for c in mapped_codes},
        missing=list(missing),
    )


# ── 1. 빈입력 → 노동·인권 섹션 전 문항이 체크리스트에 구체 문서와 함께 노출 ──
def test_labor_section_checklist_lists_concrete_documents():
    sheet = build_response_sheet("saq5")
    items = build_checklist(sheet, section=SECTION)
    qids = {it.qid for it in items}
    assert {"SAQ-L-1", "SAQ-L-2", "SAQ-L-3"} <= qids

    by_id = {it.qid: it for it in items}
    # 인권정책(S-5-1) → 인권정책서/인권헌장 업로드 안내
    hr = by_id["SAQ-L-3"]
    assert hr.action == "증빙 업로드"
    assert any("인권" in doc for doc in hr.evidence_needed)
    # 결사의 자유(S-2-6) → 단체협약/노사협의회
    union = by_id["SAQ-L-2"]
    assert any("단체협약" in doc or "노사협의회" in doc for doc in union.evidence_needed)


# ── 2. 증빙 제공 시 해당 항목이 해소되어 체크리스트에서 빠진다 ──
def test_resolved_item_drops_out_of_checklist():
    # 인권정책(S-5-1)에 공시 근거가 잡히면 verified/self_reported → 체크리스트 제외
    extraction = _extraction(mapped_codes=["S-5-1"])
    sheet = build_response_sheet("saq5", extraction=extraction)

    hr_answer = next(a for a in sheet.answers if a.qid == "SAQ-L-3")
    assert hr_answer.status in ("verified", "self_reported")

    remaining = {it.qid for it in build_checklist(sheet, section=SECTION)}
    assert "SAQ-L-3" not in remaining          # 풀렸으니 빠짐
    assert "SAQ-L-2" in remaining              # 결사의 자유는 여전히 미해소


# ── 3. evidence_needed가 answer에 실려 자기기술적 ──
def test_answer_carries_evidence_needed():
    sheet = build_response_sheet("saq5")
    hr = next(a for a in sheet.answers if a.qid == "SAQ-L-3")
    assert hr.evidence_needed
    # to_dict 직렬화에도 포함
    assert "evidence_needed" in hr.to_dict()


# ── 4. exporter: 증빙 체크리스트 시트가 생성되고 노동 문항을 담는다 ──
def test_excel_has_checklist_sheet(tmp_path: Path):
    sheet = build_response_sheet("saq5", corp_name="한울정밀")
    path = export_response_sheet(sheet, tmp_path)
    wb = load_workbook(path)
    assert "증빙 체크리스트" in wb.sheetnames
    ws = wb["증빙 체크리스트"]
    rows = list(ws.iter_rows(min_row=5, values_only=True))
    questions = [str(r[2]) for r in rows]
    assert any("인권" in q for q in questions)


# ── 5. checklist_rows(평탄화)도 노동 섹션 한정 조회 가능 ──
def test_checklist_rows_section_filter():
    sheet = build_response_sheet("saq5")
    rows = checklist_rows(sheet, section=SECTION)
    assert rows
    assert all(r["섹션"] == SECTION for r in rows)
    assert all({"할 일", "올릴 문서 / 작성 사항", "안내"} <= set(r) for r in rows)
