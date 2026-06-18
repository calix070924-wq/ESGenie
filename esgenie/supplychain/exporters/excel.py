"""ResponseSheet → .xlsx (OEM 제출본).

협력사가 대기업에 제출하는 자가진단 응답서. 각 행 = 문항 1개,
답변 옆에 신뢰 배지·근거·플래그를 함께 실어 '증빙 연결된 응답'임을 드러낸다.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from ..schema import ResponseSheet

_HEADER = ["문항 ID", "섹션", "문항", "답변", "신뢰", "근거 / 비고"]


def _fmt_value(value: Any) -> str:
    if value is None:
        return "—"
    if isinstance(value, bool):
        return "예" if value else "아니오"
    if isinstance(value, list):
        return ", ".join(str(v) for v in value) if value else "—"
    return str(value)


def _fmt_evidence(answer) -> str:
    parts: list[str] = []
    for e in answer.evidence_links:
        loc = ""
        if e.page is not None:
            loc = f" p.{e.page + 1}"
        if e.bbox:
            loc += f" bbox{[round(x, 3) for x in e.bbox]}"
        parts.append(f"{e.file_name}{loc}".strip())
    ev = " / ".join(parts)
    rationale = answer.rationale
    if ev and rationale:
        return f"{rationale}\n근거: {ev}"
    return ev or rationale


def _issb_followup_rows(answers) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for answer in answers:
        issue = ""
        remediation = ""
        for flag in getattr(answer, "flags", []) or []:
            if flag.startswith("ISSB "):
                issue = flag
            elif flag.startswith("보완 증빙: "):
                remediation = flag.removeprefix("보완 증빙: ").strip()
        if not issue or not remediation:
            continue
        rows.append({
            "문항": answer.question_text,
            "ISSB 이슈": issue,
            "권장 증빙": remediation,
            "비고": answer.rationale or "",
        })
    return rows


def export_response_sheet(sheet: ResponseSheet, out_dir: str | Path) -> str:
    """응답서를 xlsx로 저장하고 경로를 반환한다."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"실사응답서_{sheet.framework_key}_{sheet.corp_name or 'corp'}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "응답서"

    # ── 제목/요약 ──
    ws["A1"] = f"{sheet.framework_label}"
    ws["A1"].font = Font(size=13, bold=True)
    ws["A2"] = (
        f"기업: {sheet.corp_name or '—'}  |  자동응답 커버리지: {sheet.coverage_pct}%  "
        f"|  검토필요: {sheet.flagged_count}건"
    )
    ws["A2"].font = Font(size=10, color="555555")

    # ── 헤더 ──
    header_row = 4
    fill = PatternFill("solid", fgColor="1F4E78")
    for col, name in enumerate(_HEADER, start=1):
        c = ws.cell(row=header_row, column=col, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    # ── 데이터 행 ──
    status_fill = {
        "verified":     PatternFill("solid", fgColor="E2EFDA"),
        "self_reported": PatternFill("solid", fgColor="FFF2CC"),
        "insufficient": PatternFill("solid", fgColor="F2F2F2"),
        "flagged":      PatternFill("solid", fgColor="FCE4E4"),
    }
    r = header_row + 1
    for a in sheet.answers:
        ws.cell(row=r, column=1, value=a.qid)
        ws.cell(row=r, column=2, value=a.section)
        ws.cell(row=r, column=3, value=a.question_text).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=4, value=_fmt_value(a.value)).alignment = Alignment(wrap_text=True)
        badge = ws.cell(row=r, column=5, value=a.badge)
        badge.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=6, value=_fmt_evidence(a)).alignment = Alignment(wrap_text=True)
        f = status_fill.get(a.status)
        if f:
            for col in range(1, len(_HEADER) + 1):
                ws.cell(row=r, column=col).fill = f
        r += 1

    # ── 보완/검토 목록 시트 ──
    if sheet.gaps:
        gw = wb.create_sheet("보완·검토")
        gw["A1"] = "제출 전 보완·검토 항목"
        gw["A1"].font = Font(size=12, bold=True)
        for i, g in enumerate(sheet.gaps, start=3):
            gw.cell(row=i, column=1, value=g).alignment = Alignment(wrap_text=True)
        gw.column_dimensions["A"].width = 90

    issb_rows = _issb_followup_rows(sheet.answers)
    if issb_rows:
        iw = wb.create_sheet("ISSB 보완")
        iw["A1"] = "ISSB/KSSB 보완 항목"
        iw["A1"].font = Font(size=12, bold=True)
        iw["A2"] = "실사 응답서에서 ISSB 기후·그린워싱 방어 관점으로 추가 보완이 필요한 항목"
        iw["A2"].font = Font(size=10, color="555555")
        headers = ["문항", "ISSB 이슈", "권장 증빙", "비고"]
        fill = PatternFill("solid", fgColor="2E7D32")
        for col, name in enumerate(headers, start=1):
            c = iw.cell(row=4, column=col, value=name)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = fill
            c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        for row_idx, row in enumerate(issb_rows, start=5):
            iw.cell(row=row_idx, column=1, value=row["문항"]).alignment = Alignment(wrap_text=True)
            iw.cell(row=row_idx, column=2, value=row["ISSB 이슈"]).alignment = Alignment(wrap_text=True)
            iw.cell(row=row_idx, column=3, value=row["권장 증빙"]).alignment = Alignment(wrap_text=True)
            iw.cell(row=row_idx, column=4, value=row["비고"]).alignment = Alignment(wrap_text=True)
        for col, width in enumerate((48, 54, 54, 60), start=1):
            iw.column_dimensions[iw.cell(row=4, column=col).column_letter].width = width

    widths = [14, 14, 50, 24, 12, 60]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=header_row, column=col).column_letter].width = w

    wb.save(out_path)
    return str(out_path)
