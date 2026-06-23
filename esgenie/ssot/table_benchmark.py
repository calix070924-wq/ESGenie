"""표 게이트 평가셋 스캐폴드 + 리뷰 워크북 유틸리티."""
from __future__ import annotations

import json
import re
from dataclasses import asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from .ocr_router import ExtractedTable, OcrExtraction


CASE_HEADERS = [
    "id",
    "split",
    "source_file",
    "source_path",
    "doc_type",
    "page",
    "table_id",
    "gate_status_pred",
    "resolved_tier_pred",
    "needs_semantic_check_pred",
    "hard_fail_codes",
    "soft_flag_codes",
    "error_tags",
    "n_rows",
    "n_cols",
    "snapshot_path",
    "gold_table_path",
    "gate_gold",
    "review_status",
    "notes",
]

GATE_GOLD_OPTIONS = ("accept", "escalate", "human")
REVIEW_STATUS_OPTIONS = ("pending", "in_progress", "done")


def _slug(value: str) -> str:
    cleaned = re.sub(r"[^0-9A-Za-z가-힣]+", "_", value).strip("_")
    return cleaned or "table"


def make_case_id(source_file: str, page: int | None, table_id: str) -> str:
    return f"{_slug(Path(source_file).stem)}__p{(page or 0) + 1}__{_slug(table_id)}"


def serialize_table(table: ExtractedTable) -> dict[str, Any]:
    return asdict(table)


def _snapshot_payload(ext: OcrExtraction, table: ExtractedTable, case: dict[str, Any]) -> dict[str, Any]:
    table_gate = dict(ext.router_meta.get("table_gate") or {})
    payload = {
        "case": case,
        "source_file": ext.source_file,
        "doc_type": ext.doc_type,
        "channel": getattr(ext.channel, "value", str(ext.channel)),
        "router_meta": ext.router_meta,
        "table_gate": table_gate,
        "table": serialize_table(table),
        "metrics": [asdict(metric) for metric in (ext.metrics or [])],
    }
    raw_text = (ext.raw_text or "").strip()
    if raw_text:
        payload["raw_text_preview"] = raw_text[:2000]
    return payload


def extract_table_cases(
    extractions: list[OcrExtraction],
    *,
    split: str,
    raw_dir: str | Path | None = None,
    source_paths: dict[str, str] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """OcrExtraction 목록에서 표 평가셋 케이스를 만든다."""
    raw_root = Path(raw_dir) if raw_dir else None
    if raw_root:
        raw_root.mkdir(parents=True, exist_ok=True)

    cases: list[dict[str, Any]] = []
    skipped_files = 0
    skipped_sources: list[str] = []
    for ext in extractions:
        tables = list(getattr(ext, "tables", []) or [])
        if not tables:
            skipped_files += 1
            skipped_sources.append(ext.source_file)
            continue
        table_gate = dict(ext.router_meta.get("table_gate") or {})
        gate_tables = {
            row.get("table_id"): row
            for row in list(table_gate.get("tables") or [])
            if row.get("table_id")
        }
        for table in tables:
            result = gate_tables.get(table.table_id, {})
            meta = dict(getattr(table, "meta", {}) or {})
            case_id = make_case_id(ext.source_file, table.page, table.table_id)
            snapshot_path = ""
            if raw_root:
                snapshot_rel = Path("raw") / f"{case_id}.json"
                snapshot_abs = raw_root / f"{case_id}.json"
                snapshot_path = str(snapshot_rel)
            else:
                snapshot_abs = None
            case = {
                "id": case_id,
                "split": split,
                "source_file": ext.source_file,
                "source_path": (source_paths or {}).get(ext.source_file, ""),
                "doc_type": ext.doc_type,
                "page": (table.page or 0) + 1 if table.page is not None else "",
                "table_id": table.table_id,
                "gate_status_pred": result.get("decision") or table_gate.get("status") or "",
                "resolved_tier_pred": meta.get("resolved_tier", table_gate.get("resolved_tier", "")),
                "needs_semantic_check_pred": table_gate.get("needs_semantic_check", False),
                "hard_fail_codes": ";".join(meta.get("hard_fail_codes", [])),
                "soft_flag_codes": ";".join(meta.get("gate_flags", [])),
                "error_tags": ";".join(
                    sorted(
                        set(meta.get("hard_fail_codes", [])) |
                        set(meta.get("gate_flags", []))
                    )
                ),
                "n_rows": table.row_count,
                "n_cols": table.column_count,
                "snapshot_path": snapshot_path,
                "gold_table_path": f"data/benchmark_tables/gold/{case_id}.csv",
                "gate_gold": "",
                "review_status": "pending",
                "notes": "",
            }
            cases.append(case)
            if snapshot_abs is not None:
                snapshot_abs.write_text(
                    json.dumps(_snapshot_payload(ext, table, case), ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )

    summary = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "n_extractions": len(extractions),
        "n_files_without_tables": skipped_files,
        "files_without_tables": sorted(skipped_sources),
        "n_cases": len(cases),
        "by_gate_status_pred": _count_by(cases, "gate_status_pred"),
        "by_doc_type": _count_by(cases, "doc_type"),
    }
    return cases, summary


def write_cases_jsonl(cases: list[dict[str, Any]], out_path: str | Path) -> Path:
    path = Path(out_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as fp:
        for case in cases:
            fp.write(json.dumps(case, ensure_ascii=False) + "\n")
    return path


def load_cases_jsonl(path: str | Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with Path(path).open(encoding="utf-8") as fp:
        for line in fp:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def write_summary_json(summary: dict[str, Any], out_path: str | Path) -> Path:
    path = Path(out_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def build_review_workbook(cases: list[dict[str, Any]], out_path: str | Path) -> Path:
    """리뷰/라벨링용 xlsx 생성. artifact-tool 미가용 환경에서는 openpyxl 폴백."""
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cases"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cidx, name in enumerate(CASE_HEADERS, start=1):
        cell = ws.cell(row=1, column=cidx, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for ridx, row in enumerate(cases, start=2):
        for cidx, key in enumerate(CASE_HEADERS, start=1):
            value = row.get(key, "")
            if isinstance(value, bool):
                value = "TRUE" if value else "FALSE"
            ws.cell(row=ridx, column=cidx, value=value)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:T{max(2, len(cases) + 1)}"
    widths = {
        "A": 34, "B": 10, "C": 34, "D": 54, "E": 18, "F": 8, "G": 24, "H": 16, "I": 16, "J": 22,
        "K": 24, "L": 24, "M": 28, "N": 10, "O": 10, "P": 28, "Q": 28, "R": 14, "S": 14, "T": 40,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2, max_row=max(2, len(cases) + 1), min_col=1, max_col=len(CASE_HEADERS)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    dv_gate = DataValidation(type="list", formula1='"accept,escalate,human"', allow_blank=True)
    dv_review = DataValidation(type="list", formula1='"pending,in_progress,done"', allow_blank=True)
    ws.add_data_validation(dv_gate)
    ws.add_data_validation(dv_review)
    if len(cases) >= 1:
        dv_gate.add(f"R2:R{len(cases) + 1}")
        dv_review.add(f"S2:S{len(cases) + 1}")

    summary = wb.create_sheet("Summary")
    summary["A1"] = "Table Gate Review Summary"
    summary["A1"].font = Font(size=14, bold=True)
    summary["A3"] = "Metric"
    summary["B3"] = "Value"
    summary["A4"] = "Total cases"
    summary["B4"] = '=COUNTA(Cases!A:A)-1'
    summary["A5"] = "Pred ACCEPT"
    summary["B5"] = '=COUNTIF(Cases!H:H,"ACCEPT")'
    summary["A6"] = "Pred ESCALATE"
    summary["B6"] = '=COUNTIF(Cases!H:H,"ESCALATE")'
    summary["A7"] = "Pred HUMAN"
    summary["B7"] = '=COUNTIF(Cases!H:H,"HUMAN")'
    summary["A8"] = "Reviewed"
    summary["B8"] = '=COUNTIF(Cases!S:S,"done")'
    summary["A9"] = "Needs semantic check"
    summary["B9"] = '=COUNTIF(Cases!J:J,"TRUE")'
    for cell in ("A3", "B3"):
        summary[cell].font = Font(bold=True, color="FFFFFF")
        summary[cell].fill = header_fill
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 12

    chart = BarChart()
    chart.title = "Predicted Gate Decisions"
    chart.y_axis.title = "Cases"
    chart.x_axis.title = "Decision"
    data = Reference(summary, min_col=2, min_row=5, max_row=7)
    cats = Reference(summary, min_col=1, min_row=5, max_row=7)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 11
    summary.add_chart(chart, "D4")

    guide = wb.create_sheet("Guide")
    guide["A1"] = "Review instructions"
    guide["A1"].font = Font(size=14, bold=True)
    guide["A3"] = "gate_gold"
    guide["B3"] = "accept / escalate / human"
    guide["A4"] = "review_status"
    guide["B4"] = "pending / in_progress / done"
    guide["A6"] = "How to review"
    guide["A7"] = "1. snapshot_path JSON에서 원본 table/cell/bbox와 현재 gate 결과를 확인한다."
    guide["A8"] = "2. gold_table_path CSV에 최종 정답 표를 저장한다."
    guide["A9"] = "3. gate_gold는 자동 채택 허용 여부 기준으로만 라벨한다."
    guide["A10"] = "4. 애매하면 human으로 보수적으로 둔다."
    guide.column_dimensions["A"].width = 28
    guide.column_dimensions["B"].width = 72

    wb.save(out_path)
    return out_path


def _count_by(rows: list[dict[str, Any]], key: str) -> dict[str, int]:
    out: dict[str, int] = {}
    for row in rows:
        value = str(row.get(key, "") or "UNKNOWN")
        out[value] = out.get(value, 0) + 1
    return dict(sorted(out.items(), key=lambda item: item[0]))
