"""L5 — 대기업 공급망 실사 대응용 엑셀 + 증빙 서류철 생성.

산출물 2종:
  1) datasheet.xlsx — 대기업 실사 시스템에 복사·붙여넣기용 'K-ESG 정량 데이터 시트'
       · 시트1 'DataSheet' : 항목코드 | 항목명 | 값 | 단위 | 연도 | 검증상태 | D1위험 | 증빙파일
       · 시트2 'PolicyAudit': 규정 검증 결과(누락 조항 + 보완 초안 위치)
       · 시트3 'Glossary'   : 검증상태/위험도 범례
  2) evidence_pack/    — 원본 증빙 파일 복사본 (엑셀 '증빙파일' 열과 파일명 일치)
       + index.json (audit_trace_v15.json 링크)

openpyxl 사용. 증빙 파일명은 셀에 하이퍼링크로 박아 외부 감사가 클릭 추적 가능.
"""
from __future__ import annotations

import json
import shutil
from pathlib import Path
from typing import Any

from .audit_trace import AuditTraceV15, DataPoint


VERIFICATION_COLOR = {
    "verified":   "C6EFCE",   # 연녹
    "estimated":  "FFEB9C",   # 연노랑
    "unverified": "FFC7CE",   # 연빨강
}


def export_datasheet(
    trace: AuditTraceV15,
    out_dir: str | Path,
    *,
    uploaded_files: dict[str, str] | None = None,
) -> dict[str, str]:
    """엑셀 + 증빙 서류철 생성.

    Args:
        trace: build_audit_trace_v15 결과.
        out_dir: 출력 폴더(outputs/{ticker}_{ts}).
        uploaded_files: {file_name: 원본 절대경로} — evidence_pack로 복사할 소스.

    Returns: {"xlsx": ..., "audit_json": ..., "evidence_dir": ...}
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    out_dir = Path(out_dir)
    pack_dir = out_dir / "evidence_pack"
    pack_dir.mkdir(parents=True, exist_ok=True)

    # 1) 증빙 원본 복사 (엑셀 하이퍼링크 타깃)
    copied = _copy_evidence(trace, pack_dir, uploaded_files or {})

    # 2) audit_trace_v15.json 저장
    audit_path = out_dir / "audit_trace_v15.json"
    audit_path.write_text(json.dumps(trace.to_dict(), ensure_ascii=False, indent=2), "utf-8")

    # 3) 엑셀 빌드
    wb = Workbook()
    ws = wb.active
    ws.title = "DataSheet"
    headers = ["K-ESG 코드", "항목명", "값", "단위", "연도", "검증상태", "D1 위험도", "증빙 파일"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.alignment = Alignment(horizontal="center")

    for dp in trace.data_points:
        ev_name = dp.evidence_files[0].file_name if dp.evidence_files else "—"
        ws.append([
            dp.kesg_code, dp.kesg_name, dp.value, dp.unit, dp.period,
            dp.verification, dp.d1_risk, ev_name,
        ])
        row = ws.max_row
        # 검증상태 색상
        ws.cell(row, 6).fill = PatternFill(
            "solid", fgColor=VERIFICATION_COLOR.get(dp.verification, "FFFFFF"))
        # 증빙 파일 하이퍼링크
        if dp.evidence_files:
            link = dp.evidence_files[0].relative_path
            ws.cell(row, 8).hyperlink = link
            ws.cell(row, 8).font = Font(color="0563C1", underline="single")

    _autosize(ws, get_column_letter)

    # 시트2 — 규정 검증
    ws2 = wb.create_sheet("PolicyAudit")
    ws2.append(["K-ESG 코드", "통과", "요구사항", "상태", "갭 코멘트", "보완 제안"])
    for c in range(1, 7):
        ws2.cell(1, c).font = Font(bold=True)
    for pa in trace.policy_audit:
        for f in pa["findings"]:
            ws2.append([
                pa["kesg_code"], "Y" if pa["passed"] else "N",
                f.get("requirement", ""), f.get("status", ""),
                f.get("gap_comment", ""), f.get("suggested_fix", ""),
            ])
    _autosize(ws2, get_column_letter)

    # 시트3 — 범례
    ws3 = wb.create_sheet("Glossary")
    ws3.append(["검증상태", "의미"])
    ws3.append(["verified", "DART 공시 또는 증빙(±2%)으로 확인된 수치"])
    ws3.append(["estimated", "OCR 증빙 기반이나 교차검증 부분 미흡"])
    ws3.append(["unverified", "근거 노드 부재 또는 D1 위험 ≥ 0.5"])

    xlsx_path = out_dir / "ESG_DataSheet_대기업제출용.xlsx"
    wb.save(xlsx_path)

    # 4) 서류철 인덱스
    (pack_dir / "index.json").write_text(
        json.dumps({"audit_trace": "../audit_trace_v15.json", "files": copied},
                   ensure_ascii=False, indent=2), "utf-8")

    return {
        "xlsx": str(xlsx_path),
        "audit_json": str(audit_path),
        "evidence_dir": str(pack_dir),
    }


# ====================================================================
# 헬퍼
# ====================================================================

def _copy_evidence(trace: AuditTraceV15, pack_dir: Path, uploaded: dict[str, str]) -> list[str]:
    copied: list[str] = []
    for dp in trace.data_points:
        for link in dp.evidence_files:
            src = uploaded.get(link.file_name)
            if src and Path(src).exists():
                shutil.copy2(src, pack_dir / link.file_name)
                copied.append(link.file_name)
    return sorted(set(copied))


def _autosize(ws: Any, col_letter_fn: Any) -> None:
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col_letter_fn(col[0].column)].width = min(width + 4, 50)
